from django.contrib.auth import get_user_model
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken
from .models import Padre, Alumno, AlumnoMetrics, AsistenciaDetalle, Asistencia, NotificacionPadre, PersonalEducativo, Comunicado, PreferenciasPadre, Marcacion
from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, time, timedelta
import csv
import io
import os
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from config_app.models import SystemPolicy
from .services import send_attendance_email_notification, send_attendance_sms_notification
import cv2
import numpy as np


CAPTURE_MODES = {'entrada', 'salida'}


def _normalize_capture_mode(mode):
    normalized = (mode or 'entrada').strip().lower()
    return normalized if normalized in CAPTURE_MODES else 'entrada'


def _capture_tipo(mode):
    return f"facial:{_normalize_capture_mode(mode)}"


def _capture_label(tipo_marcacion):
    if not tipo_marcacion:
        return 'Puesto'

    raw_value = str(tipo_marcacion)
    if raw_value.startswith('facial:'):
        raw_value = raw_value[len('facial:'):]

    parts = raw_value.split(':', 1)
    if parts[0] in CAPTURE_MODES:
        mode = parts[0].capitalize()
        if len(parts) > 1 and parts[1]:
            return f"{mode} - {parts[1]}"
        return mode

    return raw_value or 'Puesto'


def _get_or_create_session_user(email, full_name=''):
    User = get_user_model()
    normalized_email = (email or '').strip().lower()
    if not normalized_email:
        return None

    user, created = User.objects.get_or_create(
        username=normalized_email[:150],
        defaults={'email': normalized_email},
    )
    user.email = normalized_email
    user.username = normalized_email[:150]
    if full_name:
        parts = full_name.strip().split()
        user.first_name = parts[0][:150] if parts else ''
        user.last_name = ' '.join(parts[1:])[:150] if len(parts) > 1 else ''
    if created or not user.has_usable_password():
        user.set_unusable_password()
    user.is_active = True
    user.is_staff = False
    user.is_superuser = False
    user.save()
    return user


def _get_face_match_threshold(default=75.0):
    try:
        policy = SystemPolicy.load()
        return float(policy.face_match_threshold or default)
    except Exception:
        return float(default)


class PadreLoginView(APIView):
    def post(self, request):
        try:
            email = request.data.get('email', '').strip().lower()
            dni_hijo = request.data.get('dni_hijo', '').strip()
            
            print(f"Login attempt - Email: {email}, DNI hijo: {dni_hijo}")
            
            if not email or not dni_hijo:
                return Response({'error': 'Email y DNI del hijo son requeridos'}, status=400)

            alumno = Alumno.objects.filter(email__iexact=email, dni=dni_hijo).first()
            padre = None

            if not alumno:
                padre = Padre.objects.filter(email__iexact=email).first()
                if padre:
                    alumno = Alumno.objects.filter(
                        dni=dni_hijo,
                        fk_codigo_familia=padre.fk_codigo_familia
                    ).first()

            print(f"Alumno encontrado: {alumno}")

            if not alumno:
                return Response({'error': 'Alumno no encontrado o DNI inválido'}, status=404)

            session_user = _get_or_create_session_user(
                alumno.email or email,
                f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombre}".strip(),
            )
            if not session_user:
                return Response({'error': 'No se pudo crear la sesión'}, status=500)

            refresh = RefreshToken.for_user(session_user)
            refresh['email'] = alumno.email or email
            refresh['role'] = 'parent'
            refresh['dni_hijo'] = dni_hijo
            refresh['alumno_id'] = alumno.pk_alumno
            refresh['family_code'] = str(alumno.fk_codigo_familia_id or '')

            alumno_nombre = f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}"

            payload = {
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'email': alumno.email or email,
                'nombre': alumno_nombre,
                'alumno_id': alumno.pk_alumno,
                'alumno_name': alumno_nombre,
                'family_label': f"Familia {alumno.apellido_paterno} {alumno.apellido_materno}".strip(),
                'dni_hijo': dni_hijo,
            }

            if padre:
                payload['padre_id'] = padre.pk_padre

            return Response(payload)
        except Exception as e:
            print(f"ERROR en PadreLoginView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': f'Error interno del servidor: {str(e)}'}, status=500)

class PadreAlumnosView(APIView):
    tardiness_limit = time(7, 45)

    def get(self, request, email):
        try:
            email = email.strip().lower()
            print(f"Buscando alumnos para acceso: {email}")
            alumno_por_email = Alumno.objects.filter(email__iexact=email).first()

            if alumno_por_email:
                alumnos = Alumno.objects.filter(pk_alumno=alumno_por_email.pk_alumno)
            else:
                padre = Padre.objects.filter(email__iexact=email).first()
                if not padre:
                    return Response({'error': 'Alumno o apoderado no encontrado'}, status=404)
                print(f"Padre encontrado: {padre.nombre} {padre.apellido_paterno}, Familia: {padre.fk_codigo_familia}")
                alumnos = Alumno.objects.filter(fk_codigo_familia=padre.fk_codigo_familia)
            print(f"Número de alumnos encontrados: {alumnos.count()}")
            
            alumnos_data = []
            
            for alumno in alumnos:
                print(f"Procesando alumno: {alumno.nombre} {alumno.apellido_paterno}")
                
                # Obtener asistencias con información del profesor y curso
                asistencia_qs = AsistenciaDetalle.objects.filter(
                    fk_alumno=alumno
                ).select_related(
                    'fk_asistencia', 
                    'fk_asistencia__fk_personal'
                ).order_by('-fk_asistencia__fecha')

                marcaciones_qs = Marcacion.objects.filter(
                    dni=alumno.dni,
                    tipo_marcacion__startswith='facial:'
                ).order_by('-hora_marcacion')
                
                print(f"Registros de asistencia encontrados: {asistencia_qs.count()}")
                
                attendance_by_date = {}
                
                for detalle in asistencia_qs:
                    record = self._build_attendance_from_detail(detalle)
                    if record['date']:
                        attendance_by_date[record['date']] = record
                        print(
                            f"  Registro asistencia: fecha={record['date']}, "
                            f"estado={detalle.estado_asistencia}, status={record['status']}, "
                            f"entrada={record['time_entrada']}, salida={record['time_salida']}"
                        )

                marcaciones_por_fecha = {}
                for marcacion in marcaciones_qs:
                    marcacion_date = timezone.localtime(marcacion.hora_marcacion).date().isoformat()
                    current_mark = marcaciones_por_fecha.get(marcacion_date)
                    if not current_mark or marcacion.hora_marcacion < current_mark.hora_marcacion:
                        marcaciones_por_fecha[marcacion_date] = marcacion

                for fecha_str, marcacion in marcaciones_por_fecha.items():
                    facial_record = self._build_attendance_from_marcacion(marcacion)
                    existing = attendance_by_date.get(fecha_str)

                    if not existing:
                        attendance_by_date[fecha_str] = facial_record
                        continue

                    if existing['status'] == 'absent' or not existing.get('time_entrada'):
                        merged_observations = existing.get('observations') or ''
                        if facial_record.get('observations'):
                            merged_observations = (
                                f"{merged_observations} | {facial_record['observations']}"
                                if merged_observations else facial_record['observations']
                            )
                        existing.update({
                            'status': facial_record['status'],
                            'time': facial_record['time'],
                            'time_entrada': facial_record['time_entrada'],
                            'teacher': existing.get('teacher') or facial_record['teacher'],
                            'course': existing.get('course') or facial_record['course'],
                            'observations': merged_observations,
                        })

                attendance = sorted(
                    attendance_by_date.values(),
                    key=lambda record: record.get('date') or '',
                    reverse=True,
                )
                
                alumnos_data.append({
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.nombre} {alumno.apellido_paterno} {alumno.apellido_materno}",
                    'grade': f"{alumno.grado} {alumno.seccion}" if alumno.grado and alumno.seccion else "Sin asignar",
                    'photo': alumno.foto_perfil or '👧',
                    'attendance': attendance
                })
            
            print(f"Total alumnos procesados: {len(alumnos_data)}")
            return Response({'alumnos': alumnos_data})
            
        except Exception as e:
            print(f"ERROR en PadreAlumnosView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def _build_attendance_from_detail(self, detalle):
        teacher_name = 'Sin asignar'
        curso = 'Sin curso'

        if detalle.fk_asistencia:
            if detalle.fk_asistencia.fk_personal:
                personal = detalle.fk_asistencia.fk_personal
                teacher_name = f"{personal.nombre} {personal.apellido_paterno}"

            if detalle.fk_asistencia.curso:
                curso = detalle.fk_asistencia.curso

        status = 'absent'
        if detalle.estado_asistencia:
            estado_lower = detalle.estado_asistencia.lower()
            if 'presente' in estado_lower or 'asistio' in estado_lower:
                status = 'present'
            elif 'tarde' in estado_lower or 'tardanza' in estado_lower:
                status = 'late'
            elif 'ausente' in estado_lower or 'falta' in estado_lower:
                status = 'absent'

        fecha_str = str(detalle.fk_asistencia.fecha) if detalle.fk_asistencia and detalle.fk_asistencia.fecha else None
        time_entrada = str(detalle.hora_entrada) if detalle.hora_entrada else None
        time_salida = str(detalle.fk_asistencia.hora_fin) if detalle.fk_asistencia and detalle.fk_asistencia.hora_fin else None

        return {
            'date': fecha_str,
            'status': status,
            'time': time_entrada,
            'time_entrada': time_entrada,
            'time_salida': time_salida,
            'teacher': teacher_name,
            'course': curso,
            'observations': detalle.observacion or '',
        }

    def _build_attendance_from_marcacion(self, marcacion):
        local_mark = timezone.localtime(marcacion.hora_marcacion)
        time_value = local_mark.strftime('%H:%M:%S')
        dispositivo = _capture_label(marcacion.tipo_marcacion)
        status = 'late' if local_mark.time() > self.tardiness_limit else 'present'

        return {
            'date': local_mark.date().isoformat(),
            'status': status,
            'time': time_value,
            'time_entrada': time_value,
            'time_salida': None,
            'teacher': 'Control facial',
            'course': 'Ingreso',
            'observations': f"Marcacion facial registrada en {dispositivo}.",
        }


class AdminEnrollmentStudentsView(APIView):
    def get(self, request):
        try:
            alumnos = Alumno.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombre')
            students_data = []

            for alumno in alumnos:
                metrics = AlumnoMetrics.objects.filter(fk_alumno=alumno).order_by('-ultima_captura').first()
                has_embedding = bool(metrics and metrics.embedding)

                if has_embedding:
                    template_status = 'complete'
                elif metrics:
                    template_status = 'incomplete'
                else:
                    template_status = 'pending'

                last_update = '--'
                if metrics and metrics.ultima_captura:
                    last_update = timezone.localtime(metrics.ultima_captura).strftime('%d/%m/%Y %H:%M:%S')

                enrolled_by = '--'
                if metrics and isinstance(metrics.facial_area, dict):
                    enrolled_by = metrics.facial_area.get('enrolled_by') or '--'

                students_data.append({
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}",
                    'code': alumno.codigo_alumno,
                    'dni': alumno.dni,
                    'grade': alumno.grado or 'Sin asignar',
                    'section': alumno.seccion or '-',
                    'status': alumno.estado or 'Sin estado',
                    'templateStatus': template_status,
                    'lastUpdate': last_update,
                    'enrolledBy': enrolled_by,
                    'hasTemplate': has_embedding,
                })

            return Response({
                'students': students_data,
                'total': len(students_data),
            })
        except Exception as e:
            print(f"ERROR en AdminEnrollmentStudentsView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


class AdminEnrollmentTemplateView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    min_images = 3
    max_images = 5

    def post(self, request, alumno_id):
        try:
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)

            enrolled_by = request.data.get('enrolledBy', '').strip() or 'Administrador'
            images = request.FILES.getlist('images')
            if len(images) < self.min_images or len(images) > self.max_images:
                return Response({
                    'error': f'Se requieren entre {self.min_images} y {self.max_images} imagenes.'
                }, status=400)

            descriptors = []
            facial_areas = []
            errors = []

            for index, image in enumerate(images, start=1):
                result = self._process_image(image)
                if result['error']:
                    errors.append({
                        'index': index,
                        'message': result['error'],
                    })
                    continue

                descriptors.append(result['descriptor'])
                facial_areas.append(result['facial_area'])

            if len(descriptors) < self.min_images:
                return Response({
                    'error': 'No se pudieron procesar suficientes imagenes.',
                    'details': errors,
                }, status=400)

            embedding = np.mean(np.array(descriptors, dtype=np.float32), axis=0)
            duplicate_match = self._find_duplicate_face(embedding, alumno.pk_alumno)
            if duplicate_match:
                duplicate_alumno = duplicate_match['metrics'].fk_alumno
                return Response({
                    'error': 'Este rostro ya fue enrolado para otro alumno.',
                    'duplicate': {
                        'studentId': str(duplicate_alumno.pk_alumno),
                        'name': f"{duplicate_alumno.apellido_paterno} {duplicate_alumno.apellido_materno}, {duplicate_alumno.nombre}",
                        'code': duplicate_alumno.codigo_alumno,
                        'dni': duplicate_alumno.dni,
                        'score': round(duplicate_match['score'], 2),
                        'lastUpdate': timezone.localtime(duplicate_match['metrics'].ultima_captura).strftime('%d/%m/%Y %H:%M:%S')
                            if duplicate_match['metrics'].ultima_captura else '--',
                        'enrolledBy': duplicate_match['metrics'].facial_area.get('enrolled_by', '--')
                            if isinstance(duplicate_match['metrics'].facial_area, dict) else '--',
                    },
                    'details': [
                        {'message': f"Coincide con {duplicate_alumno.apellido_paterno} {duplicate_alumno.apellido_materno}, {duplicate_alumno.nombre}"},
                        {'message': f"Codigo: {duplicate_alumno.codigo_alumno}"},
                        {'message': f"DNI: {duplicate_alumno.dni}"},
                        {'message': f"Similitud: {round(duplicate_match['score'], 2)}%"},
                    ],
                }, status=409)

            metrics = AlumnoMetrics.objects.filter(fk_alumno=alumno).order_by('-ultima_captura').first()
            if not metrics:
                metrics = AlumnoMetrics(fk_alumno=alumno)

            metrics.embedding = embedding.round(6).tolist()
            metrics.facial_area = {
                'samples': facial_areas,
                'processed_images': len(descriptors),
                'enrolled_by': enrolled_by,
            }
            metrics.porcentaje_similitud = 100.0
            metrics.ultima_captura = timezone.now()
            metrics.save()

            return Response({
                'message': 'Plantilla facial guardada correctamente.',
                'student': {
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}",
                    'code': alumno.codigo_alumno,
                },
                'template': {
                    'id': metrics.pk_alumno_metrics,
                    'processedImages': len(descriptors),
                    'lastUpdate': timezone.localtime(metrics.ultima_captura).strftime('%d/%m/%Y %H:%M:%S'),
                    'enrolledBy': enrolled_by,
                },
                'warnings': errors,
            })
        except Exception as e:
            print(f"ERROR en AdminEnrollmentTemplateView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def delete(self, request, alumno_id):
        try:
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)

            deleted_count, _ = AlumnoMetrics.objects.filter(fk_alumno=alumno).delete()
            if deleted_count == 0:
                return Response({'error': 'El alumno no tiene enrolamiento para quitar.'}, status=404)

            return Response({
                'message': 'Enrolamiento facial quitado correctamente.',
                'student': {
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}",
                    'code': alumno.codigo_alumno,
                },
            })
        except Exception as e:
            print(f"ERROR en AdminEnrollmentTemplateView DELETE: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def _find_duplicate_face(self, embedding, current_alumno_id):
        metrics_qs = (
            AlumnoMetrics.objects
            .filter(embedding__isnull=False)
            .exclude(fk_alumno_id=current_alumno_id)
            .select_related('fk_alumno')
        )
        candidate_descriptor = np.array(embedding, dtype=np.float32)
        best_match = None

        for metrics in metrics_qs:
            try:
                stored_descriptor = np.array(metrics.embedding, dtype=np.float32)
            except (TypeError, ValueError):
                continue

            if stored_descriptor.shape != candidate_descriptor.shape:
                continue

            score = self._cosine_score(candidate_descriptor, stored_descriptor)
            if score is None:
                continue

            if not best_match or score > best_match['score']:
                best_match = {
                    'score': score,
                    'metrics': metrics,
                }

        if best_match and best_match['score'] >= _get_face_match_threshold():
            return best_match

        return None

    def _cosine_score(self, first_descriptor, second_descriptor):
        denominator = float(np.linalg.norm(first_descriptor) * np.linalg.norm(second_descriptor))
        if denominator <= 0:
            return None

        cosine_similarity = float(np.dot(first_descriptor, second_descriptor) / denominator)
        return max(0.0, min(100.0, cosine_similarity * 100.0))

    def _process_image(self, image):
        content_type = (image.content_type or '').lower()
        if content_type not in {'image/jpeg', 'image/jpg', 'image/png'}:
            return {'error': 'Formato no permitido. Use JPG o PNG.'}

        file_bytes = np.frombuffer(image.read(), np.uint8)
        frame = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if frame is None:
            return {'error': 'Archivo de imagen invalido.'}

        height, width = frame.shape[:2]
        if width < 320 or height < 240:
            return {'error': 'Resolucion muy baja. Use al menos 320x240.'}

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(80, 80))

        if len(faces) == 0:
            return {'error': 'No se detecto un rostro frontal.'}
        if len(faces) > 1:
            return {'error': 'Se detecto mas de un rostro.'}

        x, y, w, h = [int(value) for value in faces[0]]
        face = gray[y:y + h, x:x + w]
        if face.size == 0:
            return {'error': 'No se pudo recortar el rostro.'}

        blur_score = cv2.Laplacian(face, cv2.CV_64F).var()
        if blur_score < 35:
            return {'error': 'La imagen esta borrosa.'}

        brightness = float(np.mean(face))
        if brightness < 45 or brightness > 220:
            return {'error': 'La iluminacion no es adecuada.'}

        normalized_face = cv2.equalizeHist(cv2.resize(face, (16, 16)))
        descriptor = (normalized_face.astype(np.float32) / 255.0).flatten()

        return {
            'error': None,
            'descriptor': descriptor.tolist(),
            'facial_area': {
                'x': x,
                'y': y,
                'w': w,
                'h': h,
                'image_width': width,
                'image_height': height,
                'blur_score': round(float(blur_score), 2),
                'brightness': round(brightness, 2),
            },
        }


class AdminFaceMatchView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    tardiness_limit = time(7, 45)

    def post(self, request):
        try:
            image = request.FILES.get('image')
            if not image:
                return Response({'error': 'Imagen requerida.'}, status=400)

            processor = AdminEnrollmentTemplateView()
            result = processor._process_image(image)
            if result['error']:
                return Response({'error': result['error']}, status=400)

            captured_descriptor = np.array(result['descriptor'], dtype=np.float32)
            metrics_qs = AlumnoMetrics.objects.filter(embedding__isnull=False).select_related('fk_alumno')
            best_match = None

            for metrics in metrics_qs:
                try:
                    stored_descriptor = np.array(metrics.embedding, dtype=np.float32)
                except (TypeError, ValueError):
                    continue

                if stored_descriptor.shape != captured_descriptor.shape:
                    continue

                distance = float(np.linalg.norm(captured_descriptor - stored_descriptor))
                denominator = float(np.linalg.norm(captured_descriptor) * np.linalg.norm(stored_descriptor))
                if denominator <= 0:
                    continue

                cosine_similarity = float(np.dot(captured_descriptor, stored_descriptor) / denominator)
                score = max(0.0, min(100.0, cosine_similarity * 100.0))

                if not best_match or score > best_match['score']:
                    best_match = {
                        'score': score,
                        'distance': distance,
                        'metrics': metrics,
                    }

            if not best_match:
                return Response({
                    'matched': False,
                    'score': 0,
                    'message': 'No hay plantillas faciales disponibles para comparar.',
                })

            metrics = best_match['metrics']
            alumno = metrics.fk_alumno
            match_threshold = _get_face_match_threshold()
            matched = best_match['score'] >= match_threshold

            if matched:
                capture_mode = _normalize_capture_mode(request.data.get('captureMode'))
                device_label = capture_mode.capitalize()
                marcacion = Marcacion.objects.create(
                    dni=alumno.dni,
                    hora_marcacion=timezone.now(),
                    tipo_marcacion=_capture_tipo(capture_mode)[:30],
                )

                local_time = timezone.localtime(marcacion.hora_marcacion)
                event_type = 'salida' if capture_mode == 'salida' else ('tardanza' if local_time.time() > self.tardiness_limit else 'entrada')
                transaction.on_commit(
                    lambda: send_attendance_email_notification(
                        alumno=alumno,
                        event_type=event_type,
                        occurred_at=marcacion.hora_marcacion,
                        device_label=device_label,
                        extra_message='Marcacion facial verificada correctamente.',
                    )
                )
                transaction.on_commit(
                    lambda: send_attendance_sms_notification(
                        alumno=alumno,
                        event_type=event_type,
                        occurred_at=marcacion.hora_marcacion,
                        device_label=device_label,
                        extra_message='Marcacion facial verificada correctamente.',
                    )
                )

            return Response({
                'matched': matched,
                'score': round(best_match['score'], 2),
                'distance': round(best_match['distance'], 4),
                'threshold': match_threshold,
                'student': {
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}",
                    'code': alumno.codigo_alumno,
                    'dni': alumno.dni,
                    'grade': alumno.grado or '',
                    'section': alumno.seccion or '',
                } if matched else None,
                'time': timezone.localtime(timezone.now()).strftime('%H:%M:%S'),
                'eventType': 'salida' if matched and _normalize_capture_mode(request.data.get('captureMode')) == 'salida' else ('entrada' if matched else None),
                'message': 'Rostro verificado correctamente.' if matched else 'Rostro detectado, pero no supera el umbral de coincidencia.',
            })
        except Exception as e:
            print(f"ERROR en AdminFaceMatchView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


class AdminTodayCaptureRecordsView(APIView):
    def get(self, request):
        try:
            today = timezone.localdate()
            marcaciones = Marcacion.objects.filter(
                hora_marcacion__date=today,
                tipo_marcacion__startswith='facial:'
            ).order_by('-hora_marcacion')

            records = []
            for marcacion in marcaciones:
                alumno = Alumno.objects.filter(dni=marcacion.dni).first()
                student_name = f"DNI {marcacion.dni}"
                if alumno:
                    student_name = f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre} ({alumno.codigo_alumno})"
                capture_label = _capture_label(marcacion.tipo_marcacion)
                event_type = 'salida' if capture_label.lower().startswith('salida') else 'entrada'

                records.append({
                    'id': str(marcacion.pk_marcacion),
                    'time': timezone.localtime(marcacion.hora_marcacion).strftime('%H:%M:%S'),
                    'eventType': event_type,
                    'student': student_name,
                    'result': 'verified',
                    'score': 100.0,
                    'message': f"Marcacion facial de {capture_label.lower()} registrada hoy.",
                })

            return Response({
                'date': str(today),
                'records': records,
            })
        except Exception as e:
            print(f"ERROR en AdminTodayCaptureRecordsView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


class AdminAttendanceDashboardView(APIView):
    tardiness_limit = time(7, 45)

    def get(self, request):
        try:
            selected_date = self._parse_date(request.GET.get('date'))
            next_date = selected_date + timedelta(days=1)
            selected_grade = request.GET.get('grade')
            selected_section = request.GET.get('section')
            selected_device = request.GET.get('device')

            marcaciones = Marcacion.objects.filter(
                hora_marcacion__gte=selected_date,
                hora_marcacion__lt=next_date,
                tipo_marcacion__startswith='facial:',
            ).order_by('-hora_marcacion')
            if selected_device and selected_device != 'all':
                marcaciones = marcaciones.filter(tipo_marcacion=f'facial:{selected_device}')

            yesterday = selected_date - timedelta(days=1)
            yesterday_next = selected_date
            yesterday_marcaciones = Marcacion.objects.filter(
                hora_marcacion__gte=yesterday,
                hora_marcacion__lt=yesterday_next,
                tipo_marcacion__startswith='facial:',
            )
            if selected_device and selected_device != 'all':
                yesterday_marcaciones = yesterday_marcaciones.filter(tipo_marcacion=f'facial:{selected_device}')

            students_qs = Alumno.objects.all()
            if selected_grade and selected_grade != 'all':
                students_qs = students_qs.filter(grado=selected_grade)
            if selected_section and selected_section != 'all':
                students_qs = students_qs.filter(seccion=selected_section)
            allowed_dnis = set(students_qs.values_list('dni', flat=True))

            unique_present_dnis = set(marcaciones.values_list('dni', flat=True)) & allowed_dnis
            unique_yesterday_dnis = set(yesterday_marcaciones.values_list('dni', flat=True)) & allowed_dnis
            total_students = students_qs.count()
            present_today = len(unique_present_dnis)
            present_yesterday = len(unique_yesterday_dnis)
            late_today = self._count_late(unique_present_dnis, selected_date)
            late_yesterday = self._count_late(unique_yesterday_dnis, yesterday)
            absences_today = max(total_students - present_today, 0)
            absences_yesterday = max(total_students - present_yesterday, 0)

            recent_events = []
            filtered_marcaciones = [marcacion for marcacion in marcaciones if marcacion.dni in allowed_dnis]
            for marcacion in filtered_marcaciones[:10]:
                alumno = Alumno.objects.filter(dni=marcacion.dni).first()
                device = self._device_label(marcacion.tipo_marcacion)
                local_time = timezone.localtime(marcacion.hora_marcacion)
                score = 100.0
                result = 'OK' if local_time.time() <= self.tardiness_limit else 'Observado'

                recent_events.append({
                    'id': str(marcacion.pk_marcacion),
                    'time': local_time.strftime('%H:%M:%S'),
                    'student': self._student_name(alumno, marcacion.dni),
                    'code': alumno.codigo_alumno if alumno else '--',
                    'device': device,
                    'result': result,
                    'score': score,
                    'liveness': True,
                })

            return Response({
                'stats': {
                    'attendanceToday': present_today,
                    'lateArrivals': late_today,
                    'absences': absences_today,
                    'totalEvents': marcaciones.count(),
                    'trend': {
                        'attendance': self._trend_percent(present_today, present_yesterday),
                        'lateArrivals': self._trend_percent(late_today, late_yesterday),
                    'absences': self._trend_percent(absences_today, absences_yesterday),
                    },
                },
                'recentEvents': recent_events,
                'filters': {
                    'grades': self._distinct_values('grado'),
                    'sections': self._distinct_values('seccion'),
                    'devices': self._distinct_devices(),
                },
                'date': str(selected_date),
            })
        except Exception as e:
            print(f"ERROR en AdminAttendanceDashboardView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def _parse_date(self, value):
        if not value or value == 'today':
            return timezone.localdate()
        if value == 'yesterday':
            return timezone.localdate() - timedelta(days=1)

        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return timezone.localdate()

    def _count_late(self, dnis, selected_date):
        if not dnis:
            return 0

        count = 0
        for dni in dnis:
            first_mark = Marcacion.objects.filter(
                dni=dni,
                hora_marcacion__date=selected_date,
                tipo_marcacion__startswith='facial:',
            ).order_by('hora_marcacion').first()
            if first_mark and timezone.localtime(first_mark.hora_marcacion).time() > self.tardiness_limit:
                count += 1
        return count

    def _trend_percent(self, current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return round(((current - previous) / previous) * 100)

    def _distinct_values(self, field_name):
        return [
            value for value in
            Alumno.objects.exclude(**{f'{field_name}__isnull': True})
            .exclude(**{field_name: ''})
            .order_by(field_name)
            .values_list(field_name, flat=True)
            .distinct()
        ]

    def _distinct_devices(self):
        devices = []
        for value in (
            Marcacion.objects.filter(tipo_marcacion__startswith='facial:')
            .order_by('tipo_marcacion')
            .values_list('tipo_marcacion', flat=True)
            .distinct()
        ):
            devices.append(self._device_label(value))
        return devices

    def _device_label(self, tipo_marcacion):
        return _capture_label(tipo_marcacion)

    def _student_name(self, alumno, dni):
        if not alumno:
            return f'DNI {dni}'
        return f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}"


class ProfesorPanelView(APIView):
    tardiness_limit = time(7, 45)

    def get(self, request, email):
        try:
            email = email.strip().lower()
            profesor = PersonalEducativo.objects.filter(email__iexact=email).first()

            if not profesor:
                return Response({'error': 'Profesor no encontrado'}, status=404)

            salon_asignado = (profesor.salon_asignado or '').strip()
            grado = None
            seccion = None
            if '-' in salon_asignado:
                grado, seccion = [part.strip() for part in salon_asignado.split('-', 1)]

            alumnos = Alumno.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombre')
            if grado:
                alumnos = alumnos.filter(grado=grado)
            if seccion:
                alumnos = alumnos.filter(seccion=seccion)

            selected_date = self._parse_selected_date(request)
            students_data = []

            for alumno in alumnos:
                detalle = AsistenciaDetalle.objects.filter(
                    fk_alumno=alumno,
                    fk_asistencia__fk_personal=profesor,
                    fk_asistencia__fecha=selected_date,
                ).select_related('fk_asistencia').order_by('-hora_entrada', '-pk_asistencia_detalle').first()

                marcacion = Marcacion.objects.filter(
                    dni=alumno.dni,
                    hora_marcacion__date=selected_date,
                ).order_by('hora_marcacion').first()

                status = 'absent'
                time_value = None

                if detalle:
                    status = self._map_attendance_status(detalle.estado_asistencia, detalle.hora_entrada)
                    if detalle.hora_entrada:
                        time_value = str(detalle.hora_entrada)[:5]

                if marcacion:
                    marcacion_time = timezone.localtime(marcacion.hora_marcacion).time()
                    marcacion_status = 'late' if marcacion_time > self.tardiness_limit else 'present'
                    if not detalle or status == 'absent':
                        status = marcacion_status
                        time_value = marcacion_time.strftime('%H:%M')

                students_data.append({
                    'id': str(alumno.pk_alumno),
                    'name': f"{alumno.nombre} {alumno.apellido_paterno} {alumno.apellido_materno}",
                    'status': status,
                    'time': time_value,
                    'photo': '👨‍🎓' if int(alumno.pk_alumno) % 2 == 0 else '👩‍🎓',
                })

            return Response({
                'teacher': {
                    'email': profesor.email,
                    'name': f"{profesor.nombre} {profesor.apellido_paterno}",
                    'classroom': salon_asignado or 'Sin aula asignada',
                    'course': profesor.especialidad or 'Curso demo',
                },
                'classes': [salon_asignado] if salon_asignado else [],
                'date': str(selected_date),
                'students': students_data,
            })
        except Exception as e:
            print(f"ERROR en ProfesorPanelView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def _parse_selected_date(self, request):
        selected_date = request.query_params.get('date')
        if not selected_date:
            return timezone.localdate()

        try:
            return datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            return timezone.localdate()

    def _map_attendance_status(self, estado_asistencia, hora_entrada):
        if estado_asistencia:
            estado_lower = estado_asistencia.lower()
            if 'presente' in estado_lower or 'asistio' in estado_lower:
                return 'present'
            if 'tarde' in estado_lower or 'tardanza' in estado_lower:
                return 'late'
            if 'ausente' in estado_lower or 'falta' in estado_lower:
                return 'absent'

        if hora_entrada:
            return 'late' if hora_entrada > self.tardiness_limit else 'present'

        return 'absent'


class ExportarProfesorAsistenciaView(APIView):
    tardiness_limit = time(7, 45)

    def get(self, request, email):
        try:
            email = email.strip().lower()
            profesor = PersonalEducativo.objects.filter(email__iexact=email).first()
            if not profesor:
                return Response({'error': 'Profesor no encontrado'}, status=404)

            formato = request.GET.get('formato', 'excel').lower()
            start_date, end_date = self._parse_date_range(request)
            if start_date > end_date:
                return Response({'error': 'La fecha inicial no puede ser mayor que la fecha final'}, status=400)

            dates, rows = self._build_attendance_matrix(profesor, start_date, end_date)
            if formato == 'pdf':
                return self._generar_pdf(profesor, dates, rows, start_date, end_date)
            if formato in {'excel', 'xlsx'}:
                return self._generar_excel(profesor, dates, rows, start_date, end_date)

            return Response({'error': 'Formato invalido. Use: pdf o excel'}, status=400)
        except Exception as e:
            print(f"ERROR en ExportarProfesorAsistenciaView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    def _parse_date_range(self, request):
        today = timezone.localdate()
        default_start = today - timedelta(days=29)
        start_date = self._parse_date(request.GET.get('start_date'), default_start)
        end_date = self._parse_date(request.GET.get('end_date'), today)
        return start_date, end_date

    def _parse_date(self, value, fallback):
        if not value:
            return fallback
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return fallback

    def _build_attendance_matrix(self, profesor, start_date, end_date):
        salon_asignado = (profesor.salon_asignado or '').strip()
        grado = None
        seccion = None
        if '-' in salon_asignado:
            grado, seccion = [part.strip() for part in salon_asignado.split('-', 1)]

        alumnos = Alumno.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombre')
        if grado:
            alumnos = alumnos.filter(grado=grado)
        if seccion:
            alumnos = alumnos.filter(seccion=seccion)

        dates = []
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date)
            current_date += timedelta(days=1)

        rows = []
        for alumno in alumnos:
            attendance_by_date = []
            for current_date in dates:
                status, time_value = self._resolve_student_status(profesor, alumno, current_date)
                attendance_by_date.append(self._cell_label(status, time_value))

            rows.append({
                'codigo': alumno.codigo_alumno,
                'dni': alumno.dni,
                'alumno': f"{alumno.apellido_paterno} {alumno.apellido_materno}, {alumno.nombre}",
                'attendance': attendance_by_date,
            })

        return dates, rows

    def _resolve_student_status(self, profesor, alumno, selected_date):
        detalle = AsistenciaDetalle.objects.filter(
            fk_alumno=alumno,
            fk_asistencia__fk_personal=profesor,
            fk_asistencia__fecha=selected_date,
        ).select_related('fk_asistencia').order_by('-hora_entrada', '-pk_asistencia_detalle').first()

        marcacion = Marcacion.objects.filter(
            dni=alumno.dni,
            hora_marcacion__date=selected_date,
        ).order_by('hora_marcacion').first()

        status = 'absent'
        time_value = None

        if detalle:
            status = self._map_attendance_status(detalle.estado_asistencia, detalle.hora_entrada)
            if detalle.hora_entrada:
                time_value = str(detalle.hora_entrada)[:5]

        if marcacion:
            marcacion_time = timezone.localtime(marcacion.hora_marcacion).time()
            marcacion_status = 'late' if marcacion_time > self.tardiness_limit else 'present'
            if not detalle or status == 'absent':
                status = marcacion_status
                time_value = marcacion_time.strftime('%H:%M')

        return status, time_value

    def _map_attendance_status(self, estado_asistencia, hora_entrada):
        if estado_asistencia:
            estado_lower = estado_asistencia.lower()
            if 'presente' in estado_lower or 'asistio' in estado_lower:
                return 'present'
            if 'tarde' in estado_lower or 'tardanza' in estado_lower:
                return 'late'
            if 'ausente' in estado_lower or 'falta' in estado_lower:
                return 'absent'

        if hora_entrada:
            return 'late' if hora_entrada > self.tardiness_limit else 'present'
        return 'absent'

    def _status_label(self, status):
        return {
            'present': 'Presente',
            'late': 'Tardanza',
            'absent': 'Falta',
            'pending': 'Pendiente',
        }.get(status, 'Falta')

    def _cell_label(self, status, time_value):
        label = self._status_label(status)
        if time_value and status in {'present', 'late'}:
            return f"{label} {time_value}"
        return label

    def _generar_excel(self, profesor, dates, rows, start_date, end_date):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Asistencia"

        last_column = 3 + len(dates)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        title_cell = sheet['A1']
        title_cell.value = f"Asistencia {profesor.salon_asignado or 'Sin aula'}"
        title_cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet['A2'] = 'Profesor'
        sheet['B2'] = f"{profesor.nombre} {profesor.apellido_paterno}"
        sheet['A3'] = 'Rango'
        sheet['B3'] = f"{start_date} a {end_date}"

        headers = ['Codigo', 'DNI', 'Alumno'] + [current_date.strftime('%d/%m') for current_date in dates]
        header_row = 5
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_index, row in enumerate(rows, start=header_row + 1):
            values = [row['codigo'], row['dni'], row['alumno'], *row['attendance']]
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(horizontal='left' if col == 3 else 'center', vertical='center')
                cell.border = border

        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 34
        for col in range(4, last_column + 1):
            sheet.column_dimensions[sheet.cell(row=header_row, column=col).column_letter].width = 15

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="asistencia_{profesor.email}_{start_date}_{end_date}.xlsx"'
        )
        return response

    def _generar_pdf(self, profesor, dates, rows, start_date, end_date):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f"Asistencia {profesor.salon_asignado or 'Sin aula'}", styles['Heading1']),
            Paragraph(f"Profesor: {profesor.nombre} {profesor.apellido_paterno}", styles['Normal']),
            Paragraph(f"Rango: {start_date} a {end_date}", styles['Normal']),
            Spacer(1, 12),
        ]

        week_groups = self._weekday_week_groups(dates)
        for week_number, week_indexes in enumerate(week_groups, start=1):
            week_dates = [dates[index] for index in week_indexes]
            if not week_dates:
                continue

            week_label = f"Semana {week_number}: {week_dates[0].strftime('%d/%m/%Y')} - {week_dates[-1].strftime('%d/%m/%Y')}"
            elements.append(Paragraph(week_label, styles['Heading3']))

            table_data = [['Alumno', *[current_date.strftime('%a %d/%m') for current_date in week_dates]]]
            for row in rows:
                table_data.append([row['alumno'][:34], *[row['attendance'][index] for index in week_indexes]])

            available_width = landscape(A4)[0] - 36
            student_column_width = 2.6 * inch
            date_column_width = (available_width - student_column_width) / max(len(week_dates), 1)
            table = Table(
                table_data,
                colWidths=[student_column_width, *([date_column_width] * len(week_dates))],
                repeatRows=1,
            )
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 6))
            elements.append(self._build_week_summary_table(rows, week_indexes))
            if week_number < len(week_groups):
                elements.append(PageBreak())
            else:
                elements.append(Spacer(1, 12))
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="asistencia_{profesor.email}_{start_date}_{end_date}.pdf"'
        )
        return response

    def _weekday_week_groups(self, dates):
        groups = []
        current_group = []
        current_week = None

        for index, current_date in enumerate(dates):
            if current_date.weekday() >= 5:
                continue

            iso_year, iso_week, _ = current_date.isocalendar()
            week_key = (iso_year, iso_week)
            if current_week is not None and week_key != current_week:
                groups.append(current_group)
                current_group = []

            current_week = week_key
            current_group.append(index)

        if current_group:
            groups.append(current_group)

        return groups

    def _build_week_summary_table(self, rows, week_indexes):
        summary = {
            'Presente': 0,
            'Tardanza': 0,
            'Falta': 0,
            'Pendiente': 0,
        }

        for row in rows:
            for index in week_indexes:
                value = row['attendance'][index]
                if value.startswith('Presente'):
                    summary['Presente'] += 1
                elif value.startswith('Tardanza'):
                    summary['Tardanza'] += 1
                elif value.startswith('Pendiente'):
                    summary['Pendiente'] += 1
                else:
                    summary['Falta'] += 1

        total = sum(summary.values())
        table = Table(
            [
                ['Resumen semanal', 'Presentes', 'Tardanzas', 'Faltas', 'Pendientes', 'Total'],
                [
                    '',
                    str(summary['Presente']),
                    str(summary['Tardanza']),
                    str(summary['Falta']),
                    str(summary['Pendiente']),
                    str(total),
                ],
            ],
            colWidths=[1.5 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 0.8 * inch],
            hAlign='LEFT',
        )
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#DCFCE7')),
            ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#FEF3C7')),
            ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#FEE2E2')),
            ('BACKGROUND', (4, 1), (4, 1), colors.HexColor('#F3F4F6')),
        ]))
        return table


class AlumnoCalificacionesView(APIView):
    """
    Endpoint para obtener calificaciones del alumno
    Por ahora retorna datos de ejemplo, listo para integrar con tabla de calificaciones cuando exista
    """
    def get(self, request, alumno_id):
        try:
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)
            
            # TODO: Cuando se cree la tabla 'calificacion', reemplazar con datos reales
            # Ejemplo de query: Calificacion.objects.filter(fk_alumno=alumno)
            
            calificaciones = {
                'promedio_general': 16.5,
                'cursos': [
                    {'nombre': 'Matemáticas', 'nota': 17},
                    {'nombre': 'Comunicación', 'nota': 16},
                    {'nombre': 'Ciencias', 'nota': 16.8},
                    {'nombre': 'Inglés', 'nota': 15.5},
                    {'nombre': 'Historia', 'nota': 16.2}
                ]
            }
            
            return Response(calificaciones)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AlumnoComportamientoView(APIView):
    """
    Endpoint para obtener el comportamiento/conducta del alumno
    Por ahora retorna datos de ejemplo, listo para integrar con tabla de comportamiento cuando exista
    """
    def get(self, request, alumno_id):
        try:
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)
            
            # TODO: Cuando se cree la tabla 'comportamiento', reemplazar con datos reales
            # Ejemplo de query: Comportamiento.objects.filter(fk_alumno=alumno)
            
            comportamiento = {
                'conducta': 'A',
                'participacion': 'Muy Buena',
                'trabajo_equipo': 'Excelente',
                'puntualidad_tareas': 'Buena',
                'responsabilidad': 'Muy Buena',
                'respeto': 'Excelente'
            }
            
            return Response(comportamiento)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AlumnoComunicadosView(APIView):
    """
    Endpoint para obtener comunicados del alumno
    Integra con la tabla comunicado - muestra comunicados generales, del grado o de la familia
    """
    def get(self, request, alumno_id):
        try:
            print(f"Buscando comunicados para alumno_id: {alumno_id}")
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            print(f"Alumno encontrado: {alumno}")
            
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)
            
            print(f"Grado del alumno: {alumno.grado}, Codigo familia: {alumno.fk_codigo_familia}")
            
            # Obtener comunicados relevantes para el alumno:
            # 1. Comunicados para todos
            # 2. Comunicados para su grado específico
            # 3. Comunicados para su familia específica
            comunicados_query = Comunicado.objects.filter(
                Q(dirigido_a='todos') |
                Q(dirigido_a='grado_especifico', grado=alumno.grado) |
                Q(dirigido_a='familia_especifica', fk_codigo_familia=alumno.fk_codigo_familia),
                estado='activo'
            ).select_related('fk_personal').order_by('-prioridad', '-fecha_publicacion')[:10]
            
            print(f"Comunicados encontrados: {comunicados_query.count()}")
            
            comunicados = []
            for com in comunicados_query:
                fecha = None
                if com.fecha_evento:
                    fecha = com.fecha_evento.strftime('%d/%m')
                elif com.fecha_publicacion:
                    fecha = com.fecha_publicacion.strftime('%d/%m')
                
                # Mapear tipo a los valores esperados por el frontend
                tipo = com.tipo if com.tipo in ['info', 'success', 'warning'] else 'info'
                if com.tipo == 'urgent':
                    tipo = 'warning'
                
                # Construir el mensaje completo
                mensaje_completo = com.titulo
                if com.mensaje and com.mensaje != com.titulo:
                    mensaje_completo = f"{com.titulo}"
                
                comunicados.append({
                    'fecha': fecha,
                    'mensaje': mensaje_completo,
                    'tipo': tipo,
                    'prioridad': com.prioridad or 1
                })
            
            # Si no hay comunicados en la BD, mostrar ejemplos
            if not comunicados:
                print("No hay comunicados en BD, usando ejemplos")
                comunicados = [
                    {'fecha': '15/02', 'mensaje': 'Evaluación de matemáticas el viernes 18', 'tipo': 'info', 'prioridad': 2},
                    {'fecha': '12/02', 'mensaje': 'Excelente participación en feria de ciencias', 'tipo': 'success', 'prioridad': 1},
                    {'fecha': '10/02', 'mensaje': 'Reunión de padres - 20 de febrero, 6:00 PM', 'tipo': 'warning', 'prioridad': 3},
                    {'fecha': '08/02', 'mensaje': 'Festival de arte el 25 de febrero', 'tipo': 'info', 'prioridad': 1},
                    {'fecha': '05/02', 'mensaje': 'Olimpiadas deportivas - inscripciones abiertas', 'tipo': 'info', 'prioridad': 1}
                ]
            
            return Response({'comunicados': comunicados})
        except Exception as e:
            print(f"ERROR en AlumnoComunicadosView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


class ExportarAsistenciaView(APIView):
    """
    Endpoint para exportar asistencia del alumno en formatos PDF, CSV o Excel
    Incluye filtrado por período de tiempo
    """
    def get(self, request, alumno_id):
        try:
            # Obtener parámetros
            formato = request.GET.get('formato', 'pdf').lower()  # pdf, csv, excel
            periodo = request.GET.get('periodo', 'all')  # all, week, month, quarter, year
            
            # Validar alumno
            alumno = Alumno.objects.filter(pk_alumno=alumno_id).first()
            if not alumno:
                return Response({'error': 'Alumno no encontrado'}, status=404)
            
            # Calcular fecha de inicio según el período
            now = datetime.now()
            fecha_inicio = None
            
            if periodo == 'week':
                fecha_inicio = now - timedelta(days=7)
            elif periodo == 'month':
                fecha_inicio = now - timedelta(days=30)
            elif periodo == 'quarter':
                fecha_inicio = now - timedelta(days=90)
            elif periodo == 'year':
                fecha_inicio = now - timedelta(days=365)
            # Si es 'all', fecha_inicio queda None y trae todos
            
            # Obtener asistencias
            asistencia_qs = AsistenciaDetalle.objects.filter(
                fk_alumno=alumno
            ).select_related(
                'fk_asistencia', 
                'fk_asistencia__fk_personal'
            ).order_by('-fk_asistencia__fecha')
            
            # Filtrar por fecha si es necesario
            if fecha_inicio:
                asistencia_qs = asistencia_qs.filter(fk_asistencia__fecha__gte=fecha_inicio)
            
            # Preparar datos
            attendance_data = []
            for detalle in asistencia_qs:
                teacher_name = 'Sin asignar'
                if detalle.fk_asistencia and detalle.fk_asistencia.fk_personal:
                    personal = detalle.fk_asistencia.fk_personal
                    teacher_name = f"{personal.nombre} {personal.apellido_paterno}"
                
                status = 'Ausente'
                if detalle.estado_asistencia:
                    estado_lower = detalle.estado_asistencia.lower()
                    if 'presente' in estado_lower or 'asistio' in estado_lower:
                        status = 'Presente'
                    elif 'tarde' in estado_lower or 'tardanza' in estado_lower:
                        status = 'Tardanza'
                
                attendance_data.append({
                    'fecha': str(detalle.fk_asistencia.fecha) if detalle.fk_asistencia else '',
                    'estado': status,
                    'hora': str(detalle.hora_entrada) if detalle.hora_entrada else '',
                    'profesor': teacher_name,
                    'curso': detalle.fk_asistencia.curso if detalle.fk_asistencia and detalle.fk_asistencia.curso else '',
                    'observaciones': detalle.observacion or ''
                })
            
            # Calcular estadísticas
            total = len(attendance_data)
            presente = sum(1 for d in attendance_data if d['estado'] == 'Presente')
            tardanza = sum(1 for d in attendance_data if d['estado'] == 'Tardanza')
            ausente = sum(1 for d in attendance_data if d['estado'] == 'Ausente')
            
            # Información del alumno y reporte
            alumno_info = {
                'nombre': f"{alumno.nombre} {alumno.apellido_paterno} {alumno.apellido_materno}",
                'grado': f"{alumno.grado} {alumno.seccion}" if alumno.grado and alumno.seccion else "Sin asignar",
                'dni': alumno.dni or ''
            }
            
            periodo_texto = {
                'all': 'Todos los registros',
                'week': 'Última semana',
                'month': 'Último mes',
                'quarter': 'Último trimestre',
                'year': 'Último año'
            }.get(periodo, 'Período personalizado')
            
            # Generar archivo según formato
            if formato == 'pdf':
                return self._generar_pdf(alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente)
            elif formato == 'csv':
                return self._generar_csv(alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente)
            elif formato == 'excel':
                return self._generar_excel(alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente)
            else:
                return Response({'error': 'Formato no válido. Use: pdf, csv o excel'}, status=400)
                
        except Exception as e:
            print(f"ERROR en ExportarAsistenciaView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
    
    def _generar_pdf(self, alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente):
        """Genera un PDF con el reporte de asistencia"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Buscar logo
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.jpg')
        if not os.path.exists(logo_path):
            # Si no existe, buscar en otros posibles lugares
            logo_path = os.path.join(settings.BASE_DIR, '..', 'frontend', 'public', 'logo.jpg')
        
        # Crear tabla con logo y título
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=1*inch, height=1*inch)
                
                # Título
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#dc2626'),
                    alignment=1  # Center
                )
                title = Paragraph("REPORTE DE ASISTENCIA", title_style)
                
                # Tabla para alinear logo a la izquierda y título al centro
                header_data = [[logo, title]]
                header_table = Table(header_data, colWidths=[1.2*inch, 5.8*inch])
                header_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ]))
                elements.append(header_table)
                elements.append(Spacer(1, 20))
            except Exception as e:
                # Si falla la carga del logo, usar solo título
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#dc2626'),
                    spaceAfter=30,
                    alignment=1
                )
                elements.append(Paragraph("REPORTE DE ASISTENCIA", title_style))
                elements.append(Spacer(1, 12))
        else:
            # Si no hay logo, solo título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=30,
                alignment=1
            )
            elements.append(Paragraph("REPORTE DE ASISTENCIA", title_style))
            elements.append(Spacer(1, 12))
        
        # Información del alumno
        info_data = [
            ['Estudiante:', alumno_info['nombre']],
            ['Grado:', alumno_info['grado']],
            ['DNI:', alumno_info['dni']],
            ['Período:', periodo_texto],
            ['Fecha de reporte:', datetime.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Estadísticas
        stats_data = [
            ['Total de días', 'Presente', 'Tardanzas', 'Ausencias'],
            [str(total), str(presente), str(tardanza), str(ausente)]
        ]
        
        stats_table = Table(stats_data, colWidths=[1.5*inch]*4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3f4f6')),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))
        
        # Detalle de asistencias
        if attendance_data:
            elements.append(Paragraph("Detalle de Asistencias", styles['Heading2']))
            elements.append(Spacer(1, 12))
            
            detail_data = [['Fecha', 'Estado', 'Hora', 'Profesor', 'Observaciones']]
            
            for record in attendance_data:
                detail_data.append([
                    record['fecha'],
                    record['estado'],
                    record['hora'],
                    record['profesor'][:20] + '...' if len(record['profesor']) > 20 else record['profesor'],
                    record['observaciones'][:30] + '...' if len(record['observaciones']) > 30 else record['observaciones']
                ])
            
            detail_table = Table(detail_data, colWidths=[1*inch, 1*inch, 0.8*inch, 1.5*inch, 1.7*inch])
            
            # Estilos base
            table_styles = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]
            
            # Aplicar colores según el estado
            for i, record in enumerate(attendance_data, start=1):
                if record['estado'] == 'Tardanza':
                    table_styles.append(('BACKGROUND', (1, i), (1, i), colors.HexColor('#fef3c7')))
                    table_styles.append(('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#92400e')))
                elif record['estado'] == 'Ausente':
                    table_styles.append(('BACKGROUND', (1, i), (1, i), colors.HexColor('#fee2e2')))
                    table_styles.append(('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#991b1b')))
                elif record['estado'] == 'Presente':
                    table_styles.append(('BACKGROUND', (1, i), (1, i), colors.HexColor('#dcfce7')))
                    table_styles.append(('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#166534')))
            
            detail_table.setStyle(TableStyle(table_styles))
            elements.append(detail_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"asistencia_{alumno_info['nombre'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _generar_csv(self, alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente):
        """Genera un archivo CSV con el reporte de asistencia"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"asistencia_{alumno_info['nombre'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Agregar BOM para UTF-8
        response.write('\ufeff')
        
        # Usar punto y coma como delimitador para Excel en español
        writer = csv.writer(response, delimiter=';')
        
        # Encabezado con información
        writer.writerow(['REPORTE DE ASISTENCIA - Fe y Alegría'])
        writer.writerow([])
        writer.writerow(['Estudiante:', alumno_info['nombre']])
        writer.writerow(['Grado:', alumno_info['grado']])
        writer.writerow(['DNI:', alumno_info['dni']])
        writer.writerow(['Período:', periodo_texto])
        writer.writerow(['Fecha de reporte:', datetime.now().strftime('%d/%m/%Y %H:%M')])
        writer.writerow([])
        
        # Estadísticas
        writer.writerow(['ESTADÍSTICAS'])
        writer.writerow(['Total de días', 'Presente', 'Tardanzas', 'Ausencias'])
        writer.writerow([total, presente, tardanza, ausente])
        writer.writerow([])
        
        # Detalle de asistencias
        writer.writerow(['DETALLE DE ASISTENCIAS'])
        writer.writerow(['Fecha', 'Estado', 'Hora', 'Profesor', 'Curso', 'Observaciones'])
        
        for record in attendance_data:
            writer.writerow([
                record['fecha'],
                record['estado'],
                record['hora'],
                record['profesor'],
                record['curso'],
                record['observaciones']
            ])
        
        return response
    
    def _generar_excel(self, alumno_info, attendance_data, periodo_texto, total, presente, tardanza, ausente):
        """Genera un archivo Excel con el reporte de asistencia"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Asistencia"
        
        # Estilos
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        
        subheader_font = Font(name='Arial', size=11, bold=True)
        subheader_fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
        subheader_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        
        normal_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        sheet.merge_cells('A1:F1')
        cell = sheet['A1']
        cell.value = 'REPORTE DE ASISTENCIA - Fe y Alegría'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        # Información del alumno
        row = 3
        info_fields = [
            ('Estudiante:', alumno_info['nombre']),
            ('Grado:', alumno_info['grado']),
            ('DNI:', alumno_info['dni']),
            ('Período:', periodo_texto),
            ('Fecha de reporte:', datetime.now().strftime('%d/%m/%Y %H:%M'))
        ]
        
        for label, value in info_fields:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = bold_font
            sheet[f'B{row}'] = value
            sheet[f'B{row}'].font = normal_font
            row += 1
        
        # Estadísticas
        row += 2
        sheet[f'A{row}'] = 'ESTADÍSTICAS'
        sheet[f'A{row}'].font = subheader_font
        row += 1
        
        stats_headers = ['Total de días', 'Presente', 'Tardanzas', 'Ausencias']
        stats_values = [total, presente, tardanza, ausente]
        
        for col, (header, value) in enumerate(zip(stats_headers, stats_values), start=1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = subheader_font_white
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
            
            cell = sheet.cell(row=row+1, column=col)
            cell.value = value
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Detalle de asistencias
        row += 4
        sheet[f'A{row}'] = 'DETALLE DE ASISTENCIAS'
        sheet[f'A{row}'].font = subheader_font
        row += 1
        
        # Encabezados de tabla
        headers = ['Fecha', 'Estado', 'Hora', 'Profesor', 'Curso', 'Observaciones']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = subheader_font_white
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Datos de asistencia
        for record in attendance_data:
            row += 1
            data = [
                record['fecha'],
                record['estado'],
                record['hora'],
                record['profesor'],
                record['curso'],
                record['observaciones']
            ]
            
            for col, value in enumerate(data, start=1):
                cell = sheet.cell(row=row, column=col)
                cell.value = value
                cell.font = normal_font
                cell.alignment = left_alignment if col > 3 else center_alignment
                cell.border = border
        
        # Ajustar ancho de columnas
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 35
        
        # Guardar en buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"asistencia_{alumno_info['nombre'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PadrePreferenciasView(APIView):
    """
    Endpoint para gestionar preferencias del padre
    GET: Obtener preferencias actuales
    PUT: Actualizar preferencias
    """
    @staticmethod
    def _as_bool(value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "t", "yes", "y", "on", "si", "sí"}
        return bool(value)

    def _resolve_padre(self, email):
        padre = Padre.objects.filter(email__iexact=email).first()
        if padre:
            return padre

        alumno = Alumno.objects.filter(email__iexact=email).select_related('fk_codigo_familia').first()
        if not alumno or not alumno.fk_codigo_familia_id:
            return None

        return Padre.objects.filter(fk_codigo_familia=alumno.fk_codigo_familia).order_by('pk_padre').first()

    def get(self, request, email):
        try:
            email = email.strip().lower()
            padre = self._resolve_padre(email)
            if not padre:
                alumno = Alumno.objects.filter(email__iexact=email).first()
                if alumno:
                    return Response({
                        'telefono': '',
                        'email': alumno.email or email,
                        'direccion': '',
                        'notificaciones_email': True,
                        'notificaciones_sms': False,
                        'notificar_entrada': True,
                        'notificar_salida': True,
                        'notificar_tardanza': True,
                        'notificar_ausencia': True,
                        'notificar_asistencia': True,
                        'notificar_calificaciones': True,
                        'notificar_comportamiento': True,
                        'frecuencia_resumen': 'semanal',
                    })
                return Response({'error': 'Padre no encontrado'}, status=404)
            
            # Obtener o crear preferencias
            preferencias, created = PreferenciasPadre.objects.get_or_create(
                fk_padre=padre,
                defaults={
                    'telefono': padre.celular or '',
                    'notificaciones_email': True,
                    'notificaciones_sms': False,
                    'notificar_entrada': True,
                    'notificar_salida': True,
                    'notificar_tardanza': True,
                    'notificar_ausencia': True,
                    'notificar_asistencia': True,
                    'notificar_calificaciones': True,
                    'notificar_comportamiento': True,
                    'frecuencia_resumen': 'semanal'
                }
            )
            
            return Response({
                'telefono': preferencias.telefono or padre.celular or '',
                'email': padre.email,
                'direccion': preferencias.direccion or '',
                'notificaciones_email': preferencias.notificaciones_email,
                'notificaciones_sms': preferencias.notificaciones_sms,
                'notificar_entrada': preferencias.notificar_entrada,
                'notificar_salida': preferencias.notificar_salida,
                'notificar_tardanza': preferencias.notificar_tardanza,
                'notificar_ausencia': preferencias.notificar_ausencia,
                'notificar_asistencia': preferencias.notificar_asistencia,
                'notificar_calificaciones': preferencias.notificar_calificaciones,
                'notificar_comportamiento': preferencias.notificar_comportamiento,
                'frecuencia_resumen': preferencias.frecuencia_resumen
            })
        except Exception as e:
            print(f"ERROR en PadrePreferenciasView GET: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
    
    def put(self, request, email):
        try:
            email = email.strip().lower()
            padre = self._resolve_padre(email)
            if not padre:
                alumno = Alumno.objects.filter(email__iexact=email).first()
                if alumno:
                    padre = Padre.objects.create(
                        nombre=alumno.nombre[:50] or 'Apoderado',
                        apellido_paterno=alumno.apellido_paterno[:50] or 'Alumno',
                        apellido_materno=alumno.apellido_materno[:50] or '',
                        dni=alumno.dni,
                        email=email,
                        fk_codigo_familia=alumno.fk_codigo_familia,
                    )
                else:
                    return Response({'error': 'Padre no encontrado'}, status=404)
            
            # Actualizar email del padre si cambió
            new_email = request.data.get('email')
            if new_email and new_email != padre.email:
                padre.email = new_email
                padre.save()
            
            # Actualizar celular si cambió
            new_telefono = request.data.get('telefono')
            if new_telefono:
                padre.celular = new_telefono
                padre.save()
            
            # Obtener o crear preferencias
            preferencias, created = PreferenciasPadre.objects.get_or_create(fk_padre=padre)
            
            # Actualizar preferencias
            preferencias.telefono = request.data.get('telefono', preferencias.telefono)
            preferencias.direccion = request.data.get('direccion', preferencias.direccion)
            preferencias.notificaciones_email = self._as_bool(request.data.get('notificaciones_email'), preferencias.notificaciones_email)
            preferencias.notificaciones_sms = self._as_bool(request.data.get('notificaciones_sms'), preferencias.notificaciones_sms)
            preferencias.notificar_entrada = self._as_bool(request.data.get('notificar_entrada'), preferencias.notificar_entrada)
            preferencias.notificar_salida = self._as_bool(request.data.get('notificar_salida'), preferencias.notificar_salida)
            preferencias.notificar_tardanza = self._as_bool(request.data.get('notificar_tardanza'), preferencias.notificar_tardanza)
            preferencias.notificar_ausencia = self._as_bool(request.data.get('notificar_ausencia'), preferencias.notificar_ausencia)
            preferencias.notificar_asistencia = self._as_bool(request.data.get('notificar_asistencia'), preferencias.notificar_asistencia)
            preferencias.notificar_calificaciones = self._as_bool(request.data.get('notificar_calificaciones'), preferencias.notificar_calificaciones)
            preferencias.notificar_comportamiento = self._as_bool(request.data.get('notificar_comportamiento'), preferencias.notificar_comportamiento)
            preferencias.frecuencia_resumen = request.data.get('frecuencia_resumen', preferencias.frecuencia_resumen)
            
            preferencias.save()
            
            return Response({
                'message': 'Preferencias actualizadas correctamente',
                'telefono': preferencias.telefono,
                'email': padre.email,
                'direccion': preferencias.direccion,
                'notificaciones_email': preferencias.notificaciones_email,
                'notificaciones_sms': preferencias.notificaciones_sms,
                'notificar_entrada': preferencias.notificar_entrada,
                'notificar_salida': preferencias.notificar_salida,
                'notificar_tardanza': preferencias.notificar_tardanza,
                'notificar_ausencia': preferencias.notificar_ausencia,
                'notificar_asistencia': preferencias.notificar_asistencia,
                'notificar_calificaciones': preferencias.notificar_calificaciones,
                'notificar_comportamiento': preferencias.notificar_comportamiento,
                'frecuencia_resumen': preferencias.frecuencia_resumen
            })
        except Exception as e:
            print(f"ERROR en PadrePreferenciasView PUT: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
